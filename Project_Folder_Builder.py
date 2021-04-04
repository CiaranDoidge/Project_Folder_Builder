# Python program to make new project folder structure
#!/usr/bin/python3
# importing os module
import os
import shutil
import openpyxl

import tkinter as tk
from tkinter import messagebox

fields = 'Project Number', 'Project Name', 'Client Name',
global Colour
Colour = 'light grey'


def fetch(entries):
    Prj_Num = entries[0][1].get()
    Prj_Name = entries[1][1].get()
    Prj_Client = entries[2][1].get()

    FolderGenerate(Prj_Num, Prj_Name, Prj_Client)


def makeform(root, fields):
    entries = []
    for field in fields:
        row = tk.Frame(root)
        lab = tk.Label(row, width=15, text=field, anchor='w')
        ent = tk.Entry(row, bg=Colour)
        row.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
        lab.pack(side=tk.LEFT)
        ent.pack(side=tk.RIGHT, expand=tk.YES, fill=tk.X)
        entries.append((field, ent))
    return entries


def FolderGenerate(Prj_Num, Prj_Name, Prj_Client):
    # Path
    parent_dir = os.getcwd()
    dir_name = (str(Prj_Num)+' - '+str(Prj_Name))

    try:
        os.mkdir(dir_name)
    except FileExistsError:
        # directory already exists
        pass
    # Make Folder Structure
    os.chdir(dir_name)
    top_level_dir = os.getcwd()
    folders = ['BOM', 'Images', 'Master Development',
               'Master Received', 'Master Released', 'Procurement']
    for folder in folders:
        try:
            os.mkdir(folder)
        except FileExistsError:
            # directory already exists
            pass

    os.chdir('Master Development')
    MDfolders = ['Mechanical', 'Hardware', 'Embedded']
    for folder in MDfolders:
        try:
            os.mkdir(folder)
        except FileExistsError:
            # directory already exists
            pass

    # Copy BOM and Procurement register from Template Folder
    # Change to top level of project folder and get path to 'BOM' folder
    os.chdir(top_level_dir)
    os.chdir('BOM')
    dest_folder = os.getcwd()
    # Change to 'Templates folder and copy to 'BOM' folder
    os.chdir(parent_dir)
    os.chdir('Templates')
    files = [os.path.join(os.getcwd(), 'PXXX - 0A(WIP) BOM.xltx'), os.path.join(
        os.getcwd(), 'PXXX - Procurement Costs Register - 0A(WIP).xltx')]
    Newfiles = [os.path.join(dest_folder, str(Prj_Num)+' - BOM -0A(WIP).xltx'), os.path.join(
        dest_folder, str(Prj_Num)+' - Procurement Costs Register - 0A(WIP).xltx')]
    # Copy template files over top 'BOM' folder
    try:
        shutil.copy(files[0], Newfiles[0])
        shutil.copy(files[1], Newfiles[1])
    except FileExistsError:
        # directory already exists
        pass

    # Add Project number, name and client to spreadsheet
    os.chdir(top_level_dir)
    os.chdir('BOM')
    # Start by opening the spreadsheet and selecting the main sheet
    XLFile = str(Prj_Num)+' - Procurement Costs Register - 0A(WIP).xltx'
    wb = openpyxl.load_workbook(XLFile)
    sheet = wb.active
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == 'Procurement Costs':
            break
    wb.active = s
    sheet.cell(row=3, column=3, value=Prj_Num)
    sheet.cell(row=4, column=3, value=Prj_Name)
    sheet.cell(row=5, column=3, value=Prj_Client)

    wb.save(XLFile)

    # Start by opening the spreadsheet and selecting the main sheet
    XLFile = str(Prj_Num)+' - BOM -0A(WIP).xltx'
    wb = openpyxl.load_workbook(filename=XLFile)

    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == 'Overview':
            break
    wb.active = s
    sheet = wb.active
    sheet.cell(row=2, column=12, value=Prj_Num)
    sheet.cell(row=3, column=12, value=Prj_Name)
    sheet.cell(row=4, column=12, value=Prj_Client)

    wb.save(XLFile)
    tk.messagebox.showinfo(message='Project Generated Sucessfully')
    root.destroy()


if __name__ == '__main__':
    root = tk.Tk()
    root.geometry('300x150+200+40')
    root.resizable(False, False)
    root.title('Project Folder Setup')
    ents = makeform(root, fields)
    root.bind('<Return>', (lambda event, e=ents: fetch(e)))
    b1 = tk.Button(root, text='Create Project',
                   command=(lambda e=ents: fetch(e)))
    b1.pack(side=tk.LEFT, padx=5, pady=5)
    b2 = tk.Button(root, text='Cancel', command=root.quit)
    b2.pack(side=tk.LEFT, padx=5, pady=5)

    root.mainloop()
